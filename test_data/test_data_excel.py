import openpyxl

# it is pulling data from Excel file and storing as dictionary into the list in order to use in actual test
# create Test Case column with numbers (Test Case 1, Test Case 2 etc.) in Excel file to loop through it
# if you want to use specific value indicate Test Case 1 or any number in if statement (line 14)
# if you want to use all values just indicate Test Case as a general name with in operator (line 14)

class TestDataExcel:
    excel_book = openpyxl.load_workbook("C:\\pythonProject2024\\withLog\\test_data\\test_data_exc.xlsx")  # enter Excel file path to capture Excel file
    sheet = excel_book.active  # capture ACTIVE sheet (don't forget to save changes)

    data_list = []

    for i in range(1, sheet.max_row + 1):  # Loop through rows
        if "Test Case" in str(sheet.cell(row=i, column=1).value):  # Check if it's a test case row
            test_data = {}  # Create a dictionary for the test case
            for c in range(2, sheet.max_column + 1):  # Loop through columns starting from the second column
                header = sheet.cell(row=1, column=c).value  # Get the header value
                value = sheet.cell(row=i, column=c).value  # Get the corresponding value for the current row
                test_data[header] = value  # Add the key-value pair to the test_data dictionary
            data_list.append(test_data)  # Append the test_data dictionary to the data_list